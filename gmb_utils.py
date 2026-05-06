import re
from pathlib import Path
import pandas as pd

METRIC_COLUMNS = [
    "Google Search - Mobile",
    "Google Search - Desktop",
    "Google Maps - Mobile",
    "Google Maps - Desktop",
    "Calls",
    "Directions",
    "Website clicks",
]

METRIC_DESCRIPTIONS = {
    "Google Search - Mobile": "Number of people that viewed your Business Profile on Google Search using Mobile",
    "Google Search - Desktop": "Number of people that viewed your Business Profile on Google Search using Desktop",
    "Google Maps - Mobile": "Number of people that viewed your Business Profile on Google Maps using Mobile",
    "Google Maps - Desktop": "Number of people that viewed your Business Profile on Google Maps using Desktop",
    "Calls": "Number of interactions with the call button from your Business Profile",
    "Directions": "Number of requests for directions made from your Business Profile",
    "Website clicks": "Number of interactions with the website button from your Business Profile",
}

MONTH_NAMES = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]

def infer_month_label(path: Path) -> str:
    name = path.name.lower()
    for month in MONTH_NAMES:
        if month in name:
            return month.title()
    match = re.search(r"(\d{4})[-_. ]?(\d{1,2})[-_. ]?(\d{1,2})", name)
    if match:
        _, month, _ = match.groups()
        month_int = int(month)
        if 1 <= month_int <= 12:
            return MONTH_NAMES[month_int - 1].title()
    return path.stem


def normalize_columns(columns):
    normalized = []
    for col in columns:
        if isinstance(col, tuple):
            col = col[0] or col[1]
        if pd.isna(col):
            normalized.append("")
            continue
        normalized.append(str(col).strip())
    return normalized


def read_gmb_file(path_or_buffer) -> pd.DataFrame:
    raw = pd.read_csv(path_or_buffer, header=[0, 1], dtype=str)
    cols = normalize_columns(raw.columns)
    raw.columns = cols
    if "Business name" not in raw.columns:
        raw = pd.read_csv(path_or_buffer, header=0, dtype=str)
        raw.columns = normalize_columns(raw.columns)
    keep = [c for c in raw.columns if c in ("Business name", "Address") + tuple(METRIC_COLUMNS)]
    if "Business name" not in keep:
        raise ValueError("Could not find Business name column")
    df = raw[keep].copy()
    df["Business name"] = df["Business name"].astype(str).str.strip()
    for metric in METRIC_COLUMNS:
        if metric in df.columns:
            df[metric] = pd.to_numeric(df[metric].fillna(0).replace("", 0), errors="coerce").fillna(0).astype(int)
    return df


def build_comparison_table(datasets: list) -> pd.DataFrame:
    """datasets is a list of tuples: (DataFrame, label_string)"""
    if not datasets:
        return pd.DataFrame()
    
    # Determine the merge keys based on available columns
    merge_keys = ["Business name"]
    if all("Address" in d[0].columns for d in datasets):
        merge_keys.append("Address")
    
    # Create a base dataframe of all unique keys
    all_keys = pd.concat([d[0][merge_keys] for d in datasets]).drop_duplicates().reset_index(drop=True)
    output = all_keys.sort_values("Business name").reset_index(drop=True)
    
    for metric in METRIC_COLUMNS:
        for df, label in datasets:
            col_name = f"{metric} {label}"
            if metric in df.columns:
                temp = df[merge_keys + [metric]].rename(columns={metric: col_name})
                
                # To prevent duplicates from multiplying rows if there are STILL duplicate keys
                # We can group by the merge keys and sum the metric, or just drop duplicates.
                # Summing is safer:
                temp = temp.groupby(merge_keys, as_index=False)[col_name].sum()
                
                output = pd.merge(output, temp, on=merge_keys, how="left")
                output[col_name] = output[col_name].fillna(0).astype(int)

    # Filter to only the columns that actually got created, in the right order
    first_cols = merge_keys.copy()
    for metric in METRIC_COLUMNS:
        for df, label in datasets:
            col_name = f"{metric} {label}"
            if col_name in output.columns:
                first_cols.append(col_name)
                
    output = output[first_cols]
    return output


def save_comparison_excel(df: pd.DataFrame, path_or_buffer, labels: list):
    from openpyxl.styles import PatternFill, Alignment, Font
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    
    with pd.ExcelWriter(path_or_buffer, engine="openpyxl") as writer:
        # Start writing data at row 4 (0-indexed 3) and no flat header
        df.to_excel(writer, index=False, header=False, startrow=3, sheet_name="Comparison")
        workbook = writer.book
        sheet = writer.sheets["Comparison"]
        
        has_address = "Address" in df.columns
        
        # Row 1: Business name
        cell_a1 = sheet.cell(row=1, column=1, value="Business name")
        cell_a1.fill = header_fill
        cell_a1.font = bold_font
        cell_a1.alignment = center_align
        
        # Merge rows 1-3 for Business name
        sheet.merge_cells("A1:A3")
        sheet.column_dimensions["A"].width = 60
        
        current_col = 2
        
        if has_address:
            # Row 1: Address
            cell_b1 = sheet.cell(row=1, column=2, value="Address")
            cell_b1.fill = header_fill
            cell_b1.font = bold_font
            cell_b1.alignment = center_align
            
            # Merge rows 1-3 for Address
            sheet.merge_cells("B1:B3")
            sheet.column_dimensions["B"].width = 60
            current_col = 3
        
        # Build metric groupings based on current dataframe
        num_months = len(labels)
        
        for metric in METRIC_COLUMNS:
            # Check if this metric is in our df
            metric_cols = [c for c in df.columns if c.startswith(metric)]
            if not metric_cols:
                continue
                
            c_start = current_col
            c_end = current_col + num_months - 1
            
            # Row 1: Main Metric Header (yellow, merged)
            m_cell = sheet.cell(row=1, column=c_start, value=metric)
            m_cell.fill = header_fill
            m_cell.font = bold_font
            m_cell.alignment = center_align
            if c_start < c_end:
                sheet.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
            
            # Row 2: Description
            desc = METRIC_DESCRIPTIONS.get(metric, "")
            d_cell = sheet.cell(row=2, column=c_start, value=desc)
            d_cell.alignment = center_align
            if c_start < c_end:
                sheet.merge_cells(start_row=2, start_column=c_start, end_row=2, end_column=c_end)
            
            # Row 3: Sub headers (months)
            for i, current_label in enumerate(labels):
                c = c_start + i
                s_cell = sheet.cell(row=3, column=c, value=current_label)
                s_cell.font = bold_font
                s_cell.alignment = center_align
                sheet.column_dimensions[get_column_letter(c)].width = 20
            
            current_col += num_months

        # Freeze panes below the headers and business name column
        sheet.freeze_panes = "C4" if has_address else "B4"
